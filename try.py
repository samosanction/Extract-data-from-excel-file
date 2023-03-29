# TASK 1. List each Company with respective product count in the given Excel sheet
# TASK 2. get the total inventory value for each company
# TASK 2. List products with inventory less than 10
# TASK 4. Add the Total inventory calculation to a new column in the Excel file
# Save as a new Excel file
import openpyxl

workbook = openpyxl.load_workbook("inventory.xlsx")
sheet = workbook["Sheet1"]

product_count = {}
total_inventory_value = {}
inv_less_ten = {}

for product_row in range(2, sheet.max_row + 1):
    supplier_name = sheet.cell(product_row, 4).value
    inventory = sheet.cell(product_row, 2).value
    price = sheet.cell(product_row, 3).value
    product_num = sheet.cell(product_row, 1).value
    total_inventory = sheet.cell(product_row, 5)

    if supplier_name in product_count:
        current_product_count = product_count[supplier_name]
        product_count[supplier_name] = product_count[supplier_name] + 1

    else:
        product_count[supplier_name] = 1

    if supplier_name in total_inventory_value:
        current_total_inv_value = total_inventory_value[supplier_name]
        total_inventory_value[supplier_name] = total_inventory_value[supplier_name] + inventory * price
    else:
        total_inventory_value[supplier_name] = inventory * price

    if inventory < 10:
        inv_less_ten[product_num] = inventory

    total_inventory.value = inventory * price
    workbook.save("updated_inventory.xlsx")

print(product_count)
print(total_inventory_value)
print(inv_less_ten)
