# TASK 1. List each Company with respective product count in the given Excel sheet
# TASK 2. get the total inventory for each company
# TASK 2. List products with inventory less than 10
import openpyxl

workbook = openpyxl.load_workbook("inventory.xlsx")
sheet = workbook["Sheet1"]

total_products = {}
total_inventory = {}
total_inv_less_ten = {}

for product_row in range(2, sheet.max_row + 1):
    supplier_name = sheet.cell(product_row, 4).value
    inventory = sheet.cell(product_row, 2).value
    price = sheet.cell(product_row, 3).value
    product_num = sheet.cell(product_row, 1).value
    total_price = sheet.cell(product_row, 5)

    if supplier_name in total_products:
        current_total_products = total_products[supplier_name]
        total_products[supplier_name] = total_products[supplier_name] + 1
    else:
        total_products[supplier_name] = 1

    # get the total inventory of each company
    if supplier_name in total_inventory:
        current_total_inventory = total_inventory[supplier_name]
        total_inventory[supplier_name] = current_total_inventory + inventory * price
    else:
        total_inventory[supplier_name] = inventory * price

    # get products with inventory less than ten
    if inventory < 10:
        total_inv_less_ten[product_num] = inventory

    # send total inventory price to a new column in the Excel and save

    total_price.value = inventory * price
    workbook.save("updated_inventory.xlsx")

print(total_inv_less_ten)
print(total_inventory)
print(total_products)